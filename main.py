import requests
import openpyxl
from datetime import date
from bs4 import BeautifulSoup as bs

# Get website HTML and load excel file
URL = 'https://www.gov.scot/publications/coronavirus-covid-19-daily-data-for-scotland/'
page = requests.get(URL)
wb = openpyxl.load_workbook('GCC2.xlsx')
ws = wb['Sheet1']
today=date.today()

# Get table from website
soup = bs(page.content, 'html.parser')
table=soup.find('tbody')

# Find all span elements on page
spanList = soup.find_all('span')
spanTextList = []
indices=[]

# Get contents of span elements
for span in spanList:
    spanTextList.append(span.text)

# Get index of span element containing Greater Glasgow and Clyde
for i, elem in enumerate(spanTextList):
    if 'Greater Glasgow and Clyde' in elem:
        indices.append(i)

# Column indices
dateCol = 1
casesCol = 2
activeCol = 3

# Cases and active cases relative to health board name
cases = spanTextList[indices[1]+1]
active = spanTextList[indices[1]+4]

# Set current row to first empty row
for cell in ws['A']:
    if cell.value is None:
        currentRow = cell.row
        break
    else:
        currentRow = cell.row + 1

# Write data to worksheet
ws.cell(row=currentRow,column=dateCol).value = today.strftime("%d %b")
ws.cell(row=currentRow,column=casesCol).value = spanTextList[indices[1]+1]
ws.cell(row=currentRow,column=activeCol).value = spanTextList[indices[1]+4]

wb.save('GCC2.xlsx')
